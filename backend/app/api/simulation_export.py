from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from uuid import UUID
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

from app.database import get_db
from app.models.scenario import SimulationScenario
from app.models.company import Company
from app.models.auth import User, UserRole
from app.auth.dependencies import get_current_active_user
from sqlalchemy import select

router = APIRouter(prefix="/companies/{company_id}/scenarios/{scenario_id}/export", tags=["simulation-export"])

@router.get("/")
async def export_simulation(
    company_id: UUID,
    scenario_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    # Access check
    company = await db.get(Company, company_id)
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")
    if current_user.role == UserRole.BUSINESS_OWNER and company.owner_id != current_user.id:
        raise HTTPException(status_code=403, detail="Access denied")

    # Fetch scenario
    scenario = await db.get(SimulationScenario, scenario_id)
    if not scenario or scenario.company_id != company_id:
        raise HTTPException(status_code=404, detail="Scenario not found")

    wb = openpyxl.Workbook()
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="D4AF37", end_color="D4AF37", fill_type="solid")
    section_font = Font(bold=True, size=11)

    # ── Sheet 1: Summary ──
    ws = wb.active
    ws.title = "Summary"
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20

    ws.append(["Company", company.name])
    ws.append(["Scenario Name", scenario.name])
    ws.append(["Mode", scenario.mode])
    ws.append(["Notes", scenario.notes or ""])
    ws.append(["Date", scenario.created_at.isoformat()])
    ws.append([])

    # Assumptions
    ws.append(["Assumptions"])
    ws[-1][0].font = section_font
    assumptions = scenario.assumptions
    for key, value in assumptions.items():
        if isinstance(value, list):
            ws.append([key, str(value)])
        else:
            ws.append([key, value])

    ws.append([])
    ws.append(["Results"])
    ws[-1][0].font = section_font
    results = scenario.results
    if scenario.mode == "deterministic":
        ws.append(["Runway (months)", results.get("runway_months", "")])
        ws.append(["Break‑even Month", results.get("break_even_month", "")])
        ws.append(["Final Cash Balance", results.get("final_cash_balance", "")])
        ws.append(["IRR", f"{results.get('irr', '')}%"])
    else:
        ws.append(["Runway P50", results.get("runway_percentiles", {}).get("p50", "")])
        ws.append(["Runway P10", results.get("runway_percentiles", {}).get("p10", "")])
        ws.append(["Runway P90", results.get("runway_percentiles", {}).get("p90", "")])
        ws.append(["Break‑even Probability", f"{results.get('breakeven_probability', '')}%"])
        ws.append(["Cash Positive Probability", f"{results.get('cash_positive_probability', '')}%"])
        ws.append(["Final Cash P50", results.get("final_cash_percentiles", {}).get("p50", "")])
        ws.append(["IRR P50", f"{results.get('irr_p50', '')}%"])

    # ── Sheet 2: Monthly Projections ──
    ws2 = wb.create_sheet("Projections")
    projections = results.get("projections", [])
    if projections:
        # Deterministic projections
        headers = list(projections[0].keys())
        ws2.append([h.replace('_', ' ').title() for h in headers])
        for row in projections:
            ws2.append([row[h] for h in headers])
    else:
        # Monte Carlo bands
        cash_bands = results.get("cash_balance_bands", [])
        revenue_bands = results.get("revenue_bands", [])
        ws2.append(["Month", "Cash P10", "Cash P50", "Cash P90", "Revenue P10", "Revenue P50", "Revenue P90"])
        for i in range(len(cash_bands)):
            c = cash_bands[i]
            r = revenue_bands[i] if i < len(revenue_bands) else {}
            ws2.append([i+1, c.get("p10", ""), c.get("p50", ""), c.get("p90", ""),
                        r.get("p10", ""), r.get("p50", ""), r.get("p90", "")])

    # ── Sheet 3: Cash Balance Data (for charting) ──
    ws3 = wb.create_sheet("Cash Balance Data")
    if projections:
        ws3.append(["Month", "Cash Balance"])
        for p in projections:
            ws3.append([p["month"], p["cash_balance"]])
    elif cash_bands:
        ws3.append(["Month", "P10", "P50", "P90"])
        for i, c in enumerate(cash_bands):
            ws3.append([i+1, c["p10"], c["p50"], c["p90"]])

    # Style headers
    for ws_name in [ws, ws2, ws3]:
        sheet = wb[ws_name]
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    # Stream response
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"{company.name.replace(' ', '_')}_{scenario.name.replace(' ', '_')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
