from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Form
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from typing import List, Optional, Dict, Any
from uuid import UUID, uuid4
from datetime import datetime
import asyncio
from pathlib import Path
from decimal import Decimal

from app.database import get_db, async_session as global_async_session
from app.models.company import Company, Document, FinancialRecord
from app.models.analytics import Valuation, RiskAssessment
from app.models.reference import SectorBaseline, ExitMultiple, RiskAdjustmentRule
from app.models.auth import User, UserRole
from app.auth.dependencies import get_current_active_user
from app.schemas.document import DocumentOut, DocumentType, RatingsEditRequest
from app.utils.storage import save_upload, delete_file
from app.engines.valuation import scorecard_valuation, vc_method, risk_adjusted_valuation
from app.engines.risk_scoring import calculate_risk_score
from app.llm_client import llm_complete

router = APIRouter(prefix="/companies/{company_id}/documents", tags=["documents"])

# ---------- Helpers ----------
def can_manage_documents(user: User, company: Company) -> bool:
    return user.role == UserRole.ADMIN or company.owner_id == user.id

def can_view_company_docs(user: User, company: Company) -> bool:
    if user.role in (UserRole.ADMIN, UserRole.INVESTOR):
        return True
    if user.role == UserRole.BUSINESS_OWNER:
        return company.owner_id == user.id
    return False

def can_edit_ratings(user: User, company: Company) -> bool:
    return user.role in (UserRole.ADMIN, UserRole.INVESTOR)

def extract_text_from_pdf(file_path: str) -> str:
    import fitz
    text = ""
    with fitz.open(file_path) as pdf:
        for page in pdf:
            text += page.get_text()
    return text

def extract_text_from_excel(file_path: str) -> str:
    import openpyxl
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    parts = []
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        parts.append(f"--- Sheet: {sheet_name} ---")
        for row in sheet.iter_rows(values_only=True):
            row_str = "\t".join(str(cell) if cell is not None else "" for cell in row)
            if row_str.strip():
                parts.append(row_str)
    return "\n".join(parts)

# ---------- Background processing wrappers ----------
async def process_pdf_document(doc_id: str, company_id: str, file_path: str | None, text: str):
    """Wrapper that creates its own database session for PDF/funding application processing."""
    async with global_async_session() as db:
        doc = await db.get(Document, UUID(doc_id))
        if not doc:
            return
        await _process_pdf_document_impl(db, doc, file_path, text)

async def process_financial_document(doc_id: str, company_id: str, file_path: str, text: str):
    """Wrapper that creates its own database session for Excel processing."""
    async with global_async_session() as db:
        doc = await db.get(Document, UUID(doc_id))
        if not doc:
            return
        await _process_financial_document_impl(db, doc, file_path, text)

# ---------- PDF processing logic ----------
async def _process_pdf_document_impl(db: AsyncSession, doc: Document, file_path: str | None, text: str):
    """Full analysis pipeline for a pitch deck or funding application."""
    import json

    if file_path and not text:
        text = extract_text_from_pdf(file_path)

    prompt = f"""You are a financial analyst for Zimbabwean startups.
Given the following text, extract:
1. Sector (one of: FinTech, AgriTech, HealthTech, CleanTech, Logistics, EdTech, Other)
2. Stage (Pre‑seed, Seed, Series A, etc.)
3. Scorecard valuation factors (team, market_size, product_tech, competitive_environment, marketing_sales, funding_need, other) – rate each as a multiplier from 0.5 to 2.0.
4. Risk scores (policy_regulatory, currency_macro, management_governance, operational_infrastructure, market_competition) – rate each from 1 (low risk) to 10 (high risk).
5. For each factor above, provide a one‑sentence justification citing specific evidence from the text.
6. Projected annual revenue in USD (if mentioned) – else null.
7. Deal terms: funding_requirement_usd, percentage_stake_offered, buyback_proposal (if any).
8. A suggested pre‑money valuation range (low‑high) in USD, purely advisory.

Return ONLY a valid JSON object. Use the following structure:
{{
  "sector": "FinTech",
  "stage": "Seed",
  "scorecard_ratings": {{
    "team": {{"rating": 1.4, "justification": "..."}},
    ...
  }},
  "risk_scores": {{
    "policy_regulatory": {{"score": 8, "justification": "..."}},
    ...
  }},
  "projected_revenue_usd": null,
  "deal_terms": {{
    "funding_requirement_usd": 100000,
    "stake_offered_percent": 60,
    "buyback_proposal": "..."
  }},
  "valuation_range_llm": {{"low": 300000, "high": 450000}}
}}

Text:
{text[:1500]}
JSON:"""

    raw_output = await llm_complete(prompt=prompt, max_tokens=1024)
    json_str = raw_output.strip()
    if json_str.startswith("```"):
        json_str = json_str.split("```")[1]
        if json_str.startswith("json"):
            json_str = json_str[4:]

    try:
        analysis = json.loads(json_str)
    except json.JSONDecodeError:
        from json_repair import repair_json
        repaired = repair_json(json_str)
        analysis = json.loads(repaired)

    doc.parsed_data = {"llm_analysis": analysis, "raw_llm_output": raw_output}
    doc.confidence_scores = {"overall": 0.85}
    doc.parse_status = "completed"

    if analysis.get("projected_revenue_usd"):
        db.add(FinancialRecord(
            company_id=doc.company_id, document_id=doc.id,
            metric_name="projected_revenue", amount=float(analysis["projected_revenue_usd"]),
            currency="USD", usd_equivalent=float(analysis["projected_revenue_usd"]), is_estimate=True
        ))

    sector = analysis.get("sector", "Other")
    stage = analysis.get("stage", "Seed")

    baseline_q = await db.execute(
        select(SectorBaseline.baseline_usd).where(SectorBaseline.sector == sector, SectorBaseline.stage == stage)
    )
    baseline_val = baseline_q.scalar() or 500_000
    if isinstance(baseline_val, Decimal):
        baseline_val = float(baseline_val)

    scorecard_ratings = {k: v["rating"] for k, v in analysis["scorecard_ratings"].items()}
    scorecard_result = scorecard_valuation(scorecard_ratings, baseline_valuation=baseline_val)
    db.add(Valuation(company_id=doc.company_id, method="scorecard", calculated_value_usd=scorecard_result["pre_money_usd"],
        pre_money_usd=scorecard_result["pre_money_usd"],
        assumptions={"ratings": scorecard_ratings, "weights": scorecard_result["weights"], "baseline": baseline_val, "source": "llm"},
        output_details=scorecard_result, performed_by="llm"))

    if analysis.get("projected_revenue_usd"):
        exit_mult_q = await db.execute(select(ExitMultiple.multiple).where(ExitMultiple.sector == sector, ExitMultiple.stage == stage))
        exit_mult = exit_mult_q.scalar() or 3.0
        if isinstance(exit_mult, Decimal): exit_mult = float(exit_mult)
        investment = 0
        deal_terms = analysis.get("deal_terms", {})
        if deal_terms and deal_terms.get("funding_requirement_usd"):
            investment = float(deal_terms["funding_requirement_usd"])
        vc_result = vc_method(projected_revenue=float(analysis["projected_revenue_usd"]), exit_multiple=exit_mult, target_return=10, investment_amount=investment)
        db.add(Valuation(company_id=doc.company_id, method="venture_capital", calculated_value_usd=vc_result["pre_money_usd"],
            pre_money_usd=vc_result["pre_money_usd"], post_money_usd=vc_result["post_money_usd"],
            assumptions=vc_result["assumptions"], output_details=vc_result, performed_by="llm"))

    risk_scores = {k: v["score"] for k, v in analysis["risk_scores"].items()}
    risk_result = calculate_risk_score(risk_scores)
    db.add(RiskAssessment(company_id=doc.company_id, overall_score=risk_result["overall_score"],
        category_scores=risk_result, assessment_notes="Auto‑generated from LLM analysis"))

    rule_q = await db.execute(select(RiskAdjustmentRule.category, RiskAdjustmentRule.score, RiskAdjustmentRule.adjustment_pct))
    rules = rule_q.all()
    adjustment_rules: Dict[str, Dict[int, int]] = {}
    for cat, sc, pct in rules:
        if cat not in adjustment_rules: adjustment_rules[cat] = {}
        adjustment_rules[cat][sc] = pct

    risk_adj_result = risk_adjusted_valuation(baseline_val, risk_scores, risk_result["weights"], adjustment_rules=adjustment_rules if adjustment_rules else None)
    db.add(Valuation(company_id=doc.company_id, method="risk_adjusted", calculated_value_usd=risk_adj_result["adjusted_valuation_usd"],
        pre_money_usd=risk_adj_result["adjusted_valuation_usd"],
        assumptions={"baseline": baseline_val, "risk_scores": risk_scores, "weights": risk_result["weights"], "source": "llm"},
        output_details=risk_adj_result, performed_by="llm"))

    if "valuation_range_llm" in analysis:
        doc.parsed_data["valuation_range_llm"] = analysis["valuation_range_llm"]

    await db.commit()

# ---------- Excel processing logic ----------
async def _process_financial_document_impl(db: AsyncSession, doc: Document, file_path: str, text: str):
    """Run LLM extraction for a financial document and store results."""
    import re

    prompt = f"""You are an accountant analyzing a Zimbabwean SME's financial document.
Extract all identifiable financial metrics as simple key-value lines, one per line.
Use the format:
  Metric Name: amount currency [YYYY-MM-DD] [projection]
Example lines:
  Revenue: 12000 USD 2025-01
  Operating Expenses: 8000 USD 2025-01 projection
  Cash Balance: 45000 USD
  Net Profit: 4000 USD 2025-01

Return ONLY the lines. No explanations, no JSON, no markdown.

Text:
{text[:1500]}
Lines:"""

    raw_output = await llm_complete(prompt=prompt, max_tokens=512)
    metrics = []
    pattern = re.compile(r"^([^:]+):\s*([\d,]+(?:\.\d+)?)\s*(\w+)(?:\s+(\d{4}-\d{2}-\d{2}))?(?:\s+(projection))?", re.IGNORECASE)
    for line in raw_output.splitlines():
        line = line.strip()
        if not line: continue
        match = pattern.match(line)
        if match:
            metric_name = match.group(1).strip()
            amount = float(match.group(2).replace(",", ""))
            currency = match.group(3).upper()
            period_end = match.group(4) if match.group(4) else None
            is_projection = bool(match.group(5))
            metrics.append({"metric_name": metric_name, "amount": amount, "currency": currency, "period_end": period_end, "is_projection": is_projection})

    doc.parsed_data = {"llm_analysis": {"financial_metrics": metrics}, "raw_llm_output": raw_output}
    doc.confidence_scores = {"overall": 0.85}
    doc.parse_status = "completed"

    for metric in metrics:
        db.add(FinancialRecord(company_id=doc.company_id, document_id=doc.id, metric_name=metric["metric_name"],
            amount=float(metric["amount"]), currency=metric.get("currency", "USD"),
            usd_equivalent=float(metric["amount"]) if metric.get("currency", "USD") == "USD" else None,
            period_end=metric.get("period_end"), is_estimate=metric.get("is_projection", False)))
    await db.commit()

# ---------- Document CRUD ----------
@router.post("/", response_model=DocumentOut, status_code=status.HTTP_201_CREATED)
async def upload_document(
    company_id: UUID, title: str = Form(...), document_type: DocumentType = Form(...),
    file: UploadFile = File(...), db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    company = await db.get(Company, company_id)
    if not company: raise HTTPException(status_code=404, detail="Company not found")
    if not can_manage_documents(current_user, company): raise HTTPException(status_code=403, detail="Access denied")

    doc_id = uuid4()
    file_path = save_upload(file, company_id, doc_id)
    document = Document(id=doc_id, company_id=company_id, title=title, document_type=document_type, file_path=file_path, parse_status="pending")
    db.add(document)
    await db.commit()
    await db.refresh(document)

    ext = Path(file.filename).suffix.lower() if file.filename else ""
    if ext in (".xlsx", ".xls"):
        text = extract_text_from_excel(file_path)
        asyncio.create_task(process_financial_document(str(doc_id), str(company_id), file_path, text))
    elif ext == ".pdf":
        text = extract_text_from_pdf(file_path)
        asyncio.create_task(process_pdf_document(str(doc_id), str(company_id), file_path, text))

    return document

@router.get("/", response_model=List[DocumentOut])
async def list_documents(company_id: UUID, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_active_user)):
    company = await db.get(Company, company_id)
    if not company: raise HTTPException(status_code=404, detail="Company not found")
    if not can_view_company_docs(current_user, company): raise HTTPException(status_code=403, detail="Access denied")
    result = await db.execute(select(Document).where(Document.company_id == company_id))
    return result.scalars().all()

@router.get("/{document_id}", response_model=DocumentOut)
async def get_document(company_id: UUID, document_id: UUID, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_active_user)):
    document = await db.get(Document, document_id)
    if not document or document.company_id != company_id: raise HTTPException(status_code=404, detail="Document not found")
    company = await db.get(Company, company_id)
    if not can_view_company_docs(current_user, company): raise HTTPException(status_code=403, detail="Access denied")
    return document

@router.post("/{document_id}/parse", response_model=DocumentOut)
async def trigger_parse(company_id: UUID, document_id: UUID, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_active_user)):
    document = await db.get(Document, document_id)
    if not document or document.company_id != company_id: raise HTTPException(status_code=404, detail="Document not found")
    company = await db.get(Company, company_id)
    if not can_manage_documents(current_user, company): raise HTTPException(status_code=403, detail="Access denied")
    if document.parse_status not in ("pending", "failed"): raise HTTPException(status_code=400, detail=f"Document already {document.parse_status}")
    document.parse_status = "processing"
    await db.commit(); await db.refresh(document)
    return document

@router.delete("/{document_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_document(company_id: UUID, document_id: UUID, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_active_user)):
    document = await db.get(Document, document_id)
    if not document or document.company_id != company_id: raise HTTPException(status_code=404, detail="Document not found")
    company = await db.get(Company, company_id)
    if not can_manage_documents(current_user, company): raise HTTPException(status_code=403, detail="Access denied")
    delete_file(document.file_path)
    await db.delete(document); await db.commit()
    return None

# ---------- Ratings Editing ----------
@router.put("/{document_id}/ratings")
async def edit_ratings(company_id: UUID, document_id: UUID, payload: RatingsEditRequest, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_active_user)):
    document = await db.get(Document, document_id)
    if not document or document.company_id != company_id: raise HTTPException(status_code=404, detail="Document not found")
    company = await db.get(Company, company_id)
    if not company: raise HTTPException(status_code=404, detail="Company not found")
    if not can_edit_ratings(current_user, company): raise HTTPException(status_code=403, detail="Only investors and admins can edit ratings")

    if document.document_type not in ("pitch_deck", "funding_application"): raise HTTPException(status_code=400, detail="Ratings editing is only available for pitch decks and funding applications")
    if not document.parsed_data or not document.parsed_data.get("llm_analysis"): raise HTTPException(status_code=400, detail="Document hasn't been analysed yet; cannot edit ratings")

    analysis = document.parsed_data["llm_analysis"]
    sector = analysis.get("sector", company.sector or "Other")
    stage = analysis.get("stage", company.stage or "Seed")

    original_scorecard = {k: v["rating"] for k, v in analysis.get("scorecard_ratings", {}).items()}
    original_risk = {k: v["score"] for k, v in analysis.get("risk_scores", {}).items()}
    new_scorecard = {**original_scorecard, **(payload.scorecard_ratings or {})}
    new_risk = {**original_risk, **(payload.risk_scores or {})}
    justifications = payload.justifications or {}

    document.parsed_data["edited_ratings"] = {
        "scorecard_ratings": {k: {"rating": v, "justification": justifications.get(k, "")} for k, v in new_scorecard.items()},
        "risk_scores": {k: {"score": v, "justification": justifications.get(k, "")} for k, v in new_risk.items()},
        "edited_by": str(current_user.id), "edited_at": datetime.utcnow().isoformat()
    }
    await db.commit(); await db.refresh(document)

    if not payload.recalculate:
        return {"message": "Ratings saved. Recalculation skipped."}

    baseline_q = await db.execute(select(SectorBaseline.baseline_usd).where(SectorBaseline.sector == sector, SectorBaseline.stage == stage))
    baseline_val = baseline_q.scalar() or 500_000
    if isinstance(baseline_val, Decimal): baseline_val = float(baseline_val)

    scorecard_result = scorecard_valuation(new_scorecard, baseline_valuation=baseline_val)
    db.add(Valuation(company_id=company_id, method="scorecard", calculated_value_usd=scorecard_result["pre_money_usd"],
        pre_money_usd=scorecard_result["pre_money_usd"],
        assumptions={"ratings": new_scorecard, "weights": scorecard_result["weights"], "baseline": baseline_val, "source": "manual_edit"},
        output_details=scorecard_result, performed_by=str(current_user.id)))

    projected_rev = analysis.get("projected_revenue_usd")
    if projected_rev:
        exit_mult_q = await db.execute(select(ExitMultiple.multiple).where(ExitMultiple.sector == sector, ExitMultiple.stage == stage))
        exit_mult = exit_mult_q.scalar() or 3.0
        if isinstance(exit_mult, Decimal): exit_mult = float(exit_mult)
        investment = float(analysis.get("deal_terms", {}).get("funding_requirement_usd", 0))
        vc_result = vc_method(projected_revenue=float(projected_rev), exit_multiple=exit_mult, target_return=10, investment_amount=investment)
        db.add(Valuation(company_id=company_id, method="venture_capital", calculated_value_usd=vc_result["pre_money_usd"],
            pre_money_usd=vc_result["pre_money_usd"], post_money_usd=vc_result["post_money_usd"],
            assumptions=vc_result["assumptions"], output_details=vc_result, performed_by=str(current_user.id)))

    risk_result = calculate_risk_score(new_risk)
    db.add(RiskAssessment(company_id=company_id, overall_score=risk_result["overall_score"],
        category_scores=risk_result, assessment_notes=f"Ratings edited by {current_user.email}"))

    rule_q = await db.execute(select(RiskAdjustmentRule.category, RiskAdjustmentRule.score, RiskAdjustmentRule.adjustment_pct))
    rules = rule_q.all()
    adjustment_rules = {}
    for cat, sc, pct in rules:
        if cat not in adjustment_rules: adjustment_rules[cat] = {}
        adjustment_rules[cat][sc] = pct

    risk_adj_result = risk_adjusted_valuation(baseline_val, new_risk, risk_result["weights"], adjustment_rules=adjustment_rules if adjustment_rules else None)
    db.add(Valuation(company_id=company_id, method="risk_adjusted", calculated_value_usd=risk_adj_result["adjusted_valuation_usd"],
        pre_money_usd=risk_adj_result["adjusted_valuation_usd"],
        assumptions={"baseline": baseline_val, "risk_scores": new_risk, "weights": risk_result["weights"], "source": "manual_edit"},
        output_details=risk_adj_result, performed_by=str(current_user.id)))

    await db.commit()
    return {"message": "Ratings updated and valuations recalculated.",
            "new_valuations": {"scorecard": scorecard_result["pre_money_usd"], "venture_capital": vc_result["pre_money_usd"] if projected_rev else None,
                               "risk_adjusted": risk_adj_result["adjusted_valuation_usd"]}, "new_risk_score": risk_result["overall_score"]}
