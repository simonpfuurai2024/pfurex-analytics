from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from uuid import UUID
from typing import List, Optional, Dict, Any
from datetime import datetime

from fastapi.responses import StreamingResponse
import io
import openpyxl
from openpyxl.styles import Font, Alignment


from app.database import get_db
from app.models.company import Company, Document, FinancialRecord
from app.models.analytics import Valuation, RiskAssessment
from app.models.auth import User, UserRole
from app.auth.dependencies import get_current_active_user

router = APIRouter(prefix="/companies/{company_id}/dashboard", tags=["dashboard"])


def can_view_company(user: User, company: Company) -> bool:
    """Admins and investors see all; business owners only see their own."""
    if user.role in (UserRole.ADMIN, UserRole.INVESTOR):
        return True
    if user.role == UserRole.BUSINESS_OWNER:
        return company.owner_id == user.id
    return False


@router.get("/")
async def get_dashboard(
    company_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    # 1. Company check & access control
    company = await db.get(Company, company_id)
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")
    if not can_view_company(current_user, company):
        raise HTTPException(status_code=403, detail="Access denied")

    # 2. Valuations: get the latest per method
    valuation_query = (
        select(Valuation)
        .where(Valuation.company_id == company_id)
        .order_by(Valuation.created_at.desc())
    )
    val_result = await db.execute(valuation_query)
    all_valuations = val_result.scalars().all()

    # Group latest per method
    latest_valuations: Dict[str, Any] = {}
    for v in all_valuations:
        if v.method not in latest_valuations:
            latest_valuations[v.method] = {
                "id": str(v.id),
                "method": v.method,
                "pre_money_usd": v.pre_money_usd,
                "post_money_usd": v.post_money_usd,
                "calculated_value_usd": v.calculated_value_usd,
                "assumptions": v.assumptions,
                "output_details": v.output_details,
                "performed_by": v.performed_by,
                "created_at": v.created_at.isoformat()
            }

    # 3. Latest Risk Assessment
    risk_query = (
        select(RiskAssessment)
        .where(RiskAssessment.company_id == company_id)
        .order_by(RiskAssessment.assessed_at.desc())
        .limit(1)
    )
    risk_result = await db.execute(risk_query)
    latest_risk = risk_result.scalar_one_or_none()
    risk_data = None
    if latest_risk:
        risk_data = {
            "id": str(latest_risk.id),
            "overall_score": latest_risk.overall_score,
            "category_scores": latest_risk.category_scores,
            "assessment_notes": latest_risk.assessment_notes,
            "assessed_at": latest_risk.assessed_at.isoformat()
        }

    # 4. Financial Records (all, ordered by period)
    fin_query = (
        select(FinancialRecord)
        .where(FinancialRecord.company_id == company_id)
        .order_by(FinancialRecord.period_end.desc())
    )
    fin_result = await db.execute(fin_query)
    financial_records = fin_result.scalars().all()
    financials = [
        {
            "id": str(f.id),
            "metric_name": f.metric_name,
            "amount": f.amount,
            "currency": f.currency,
            "usd_equivalent": f.usd_equivalent,
            "period_end": f.period_end.isoformat() if f.period_end else None,
            "is_estimate": f.is_estimate,
            "document_id": str(f.document_id) if f.document_id else None
        }
        for f in financial_records
    ]

    # 5. Documents (all)
    doc_query = (
        select(Document)
        .where(Document.company_id == company_id)
        .order_by(Document.uploaded_at.desc())
    )
    doc_result = await db.execute(doc_query)
    documents = doc_result.scalars().all()
    docs_list = [
        {
            "id": str(d.id),
            "title": d.title,
            "document_type": d.document_type,
            "parse_status": d.parse_status,
            "uploaded_at": d.uploaded_at.isoformat(),
            "has_parsed_data": d.parsed_data is not None
        }
        for d in documents
    ]

    # 6. Assemble response
    return {
        "company": {
            "id": str(company.id),
            "name": company.name,
            "sector": company.sector,
            "stage": company.stage,
            "founded_year": company.founded_year,
            "country": company.country,
            "description": company.description,
            "primary_contact": company.primary_contact,
            "owner_id": str(company.owner_id) if company.owner_id else None,
            "created_at": company.created_at.isoformat(),
            "updated_at": company.updated_at.isoformat()
        },
        "valuations": latest_valuations,          # dict keyed by method
        "risk_assessment": risk_data,
        "financial_records": financials,
        "documents": docs_list
    }

@router.get("/report")
async def get_report(
    company_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    # Reuse dashboard logic to fetch data
    company = await db.get(Company, company_id)
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")
    if not can_view_company(current_user, company):
        raise HTTPException(status_code=403, detail="Access denied")

    # Valuations (latest per method)
    valuation_query = (
        select(Valuation)
        .where(Valuation.company_id == company_id)
        .order_by(Valuation.created_at.desc())
    )
    val_result = await db.execute(valuation_query)
    all_valuations = val_result.scalars().all()
    latest_vals: Dict[str, Any] = {}
    for v in all_valuations:
        if v.method not in latest_vals:
            latest_vals[v.method] = v

    # Latest Risk
    risk_query = (
        select(RiskAssessment)
        .where(RiskAssessment.company_id == company_id)
        .order_by(RiskAssessment.assessed_at.desc())
        .limit(1)
    )
    risk_result = await db.execute(risk_query)
    latest_risk = risk_result.scalar_one_or_none()

    # Financial Records
    fin_query = (
        select(FinancialRecord)
        .where(FinancialRecord.company_id == company_id)
        .order_by(FinancialRecord.period_end.desc())
    )
    fin_result = await db.execute(fin_query)
    financials = fin_result.scalars().all()

    # Documents
    doc_query = (
        select(Document)
        .where(Document.company_id == company_id)
        .order_by(Document.uploaded_at.desc())
    )
    doc_result = await db.execute(doc_query)
    documents = doc_result.scalars().all()

    # Build Excel workbook
    wb = openpyxl.Workbook()
    # --- Sheet 1: Company Summary ---
    ws = wb.active
    ws.title = "Company Info"
    ws["A1"] = "Company Name"
    ws["B1"] = company.name
    ws["A2"] = "Sector"
    ws["B2"] = company.sector
    ws["A3"] = "Stage"
    ws["B3"] = company.stage
    ws["A4"] = "Founded Year"
    ws["B4"] = company.founded_year
    ws["A5"] = "Country"
    ws["B5"] = company.country
    ws["A6"] = "Description"
    ws["B6"] = company.description

    # --- Sheet 2: Valuations ---
    ws_val = wb.create_sheet("Valuations")
    headers = ["Method", "Pre-Money USD", "Post-Money USD", "Calculated Value", "Performed By", "Date"]
    ws_val.append(headers)
    for method, val in latest_vals.items():
        ws_val.append([
            val.method,
            val.pre_money_usd,
            val.post_money_usd,
            val.calculated_value_usd,
            val.performed_by,
            val.created_at.isoformat()
        ])

    # --- Sheet 3: Risk Assessment ---
    ws_risk = wb.create_sheet("Risk Assessment")
    ws_risk["A1"] = "Overall Score"
    ws_risk["B1"] = latest_risk.overall_score if latest_risk else "N/A"
    if latest_risk:
        ws_risk["A3"] = "Category Breakdown"
        ws_risk.append(["Category", "Weight", "Score", "Weighted Score"])
        cat_scores = latest_risk.category_scores.get("category_breakdown", {})
        for cat, details in cat_scores.items():
            ws_risk.append([
                cat,
                details.get("weight", ""),
                details.get("raw_score", ""),
                details.get("weighted_score", "")
            ])

    # --- Sheet 4: Financial Records ---
    ws_fin = wb.create_sheet("Financials")
    ws_fin.append(["Metric", "Amount", "Currency", "USD Equivalent", "Period End", "Estimated"])
    for f in financials:
        ws_fin.append([
            f.metric_name,
            f.amount,
            f.currency,
            f.usd_equivalent,
            f.period_end.isoformat() if f.period_end else "",
            "Yes" if f.is_estimate else "No"
        ])

    # --- Sheet 5: Documents ---
    ws_doc = wb.create_sheet("Documents")
    ws_doc.append(["Title", "Type", "Status", "Uploaded"])
    for d in documents:
        ws_doc.append([d.title, d.document_type, d.parse_status, d.uploaded_at.isoformat()])

    # Stream response
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"{company.name.replace(' ', '_')}_report.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
