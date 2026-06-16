import asyncio
import json
import re
import fitz
import openpyxl
from pathlib import Path
from sqlalchemy.ext.asyncio import AsyncSession
from aiokafka import AIOKafkaConsumer
from typing import Dict, Any, List
from decimal import Decimal

from app.database import async_session
from app.models.company import Company, Document, FinancialRecord
from app.models.analytics import Valuation, RiskAssessment
from app.models.reference import SectorBaseline, ExitMultiple, RiskAdjustmentRule
from app.engines.valuation import scorecard_valuation, vc_method, risk_adjusted_valuation
from app.engines.risk_scoring import calculate_risk_score
from sqlalchemy import select

REDPANDA_BROKER = "localhost:9092"
TOPIC = "document.uploaded"


def extract_text_from_pdf(file_path: str) -> str:
    text = ""
    with fitz.open(file_path) as pdf:
        for page in pdf:
            text += page.get_text()
    return text


def extract_text_from_excel(file_path: str) -> str:
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    parts = []
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        parts.append(f"--- Sheet: {sheet_name} ---")
        for row in sheet.iter_rows(values_only=True):
            row_str = "\t".join(str(cell) if cell is not None else "" for cell in row)
            if row_str.strip():
                parts.append(row_str)
    return "\n".join(parts)


def build_analysis_prompt(text: str, document_type: str) -> str:
    if document_type in ("pitch_deck", "funding_application"):
        return f"""You are a financial analyst for Zimbabwean startups.
Given the following text, extract:
1. Sector (one of: FinTech, AgriTech, HealthTech, CleanTech, Logistics, EdTech, Other)
2. Stage (Pre‑seed, Seed, Series A, etc.)
3. Scorecard valuation factors (team, market_size, product_tech, competitive_environment, marketing_sales, funding_need, other) – rate each as a multiplier from 0.5 to 2.0.
4. Risk scores (policy_regulatory, currency_macro, management_governance, operational_infrastructure, market_competition) – rate each from 1 (low risk) to 10 (high risk).
5. For each factor above, provide a one‑sentence justification citing specific evidence from the text.
6. Projected annual revenue in USD (if mentioned) – else null.
7. Deal terms: funding_requirement_usd, percentage_stake_offered, buyback_proposal (if any).
8. A suggested pre‑money valuation range (low‑high) in USD, purely advisory.

Return ONLY a valid JSON object. Use the following structure:
{{
  "sector": "FinTech",
  "stage": "Seed",
  "scorecard_ratings": {{
    "team": {{"rating": 1.4, "justification": "..."}},
    ...
  }},
  "risk_scores": {{
    "policy_regulatory": {{"score": 8, "justification": "..."}},
    ...
  }},
  "projected_revenue_usd": null,
  "deal_terms": {{
    "funding_requirement_usd": 100000,
    "stake_offered_percent": 60,
    "buyback_proposal": "..."
  }},
  "valuation_range_llm": {{"low": 300000, "high": 450000}}
}}

Text:
{text[:3000]}
JSON:"""
    else:
        # For financial documents: simple line-by-line format
        return f"""You are an accountant analyzing a Zimbabwean SME's financial document.
Extract all identifiable financial metrics as simple key-value lines, one per line.
Use the format:
  Metric Name: amount currency [YYYY-MM-DD] [projection]
Example lines:
  Revenue: 12000 USD 2025-01
  Operating Expenses: 8000 USD 2025-01 projection
  Cash Balance: 45000 USD
  Net Profit: 4000 USD 2025-01

Return ONLY the lines. No explanations, no JSON, no markdown.

Text:
{text[:3000]}
Lines:"""


def parse_financial_lines(lines_text: str) -> List[Dict[str, Any]]:
    """Parse 'Key: value' lines into financial metric dicts."""
    metrics = []
    pattern = re.compile(
        r"^([^:]+):\s*([\d,]+(?:\.\d+)?)\s*(\w+)(?:\s+(\d{4}-\d{2}-\d{2}))?(?:\s+(projection))?",
        re.IGNORECASE
    )
    for line in lines_text.splitlines():
        line = line.strip()
        if not line:
            continue
        match = pattern.match(line)
        if match:
            metric_name = match.group(1).strip()
            amount = float(match.group(2).replace(",", ""))
            currency = match.group(3).upper()
            period_end = match.group(4) if match.group(4) else None
            is_projection = bool(match.group(5))
            metrics.append({
                "metric_name": metric_name,
                "amount": amount,
                "currency": currency,
                "period_end": period_end,
                "is_projection": is_projection
            })
    return metrics


async def process_document(msg):
    data = json.loads(msg.value.decode())
    doc_id = data["document_id"]
    company_id = data["company_id"]
    file_path = data.get("file_path")

    async with async_session() as db:
        doc = await db.get(Document, doc_id)
        if not doc or doc.parse_status != "pending":
            return

        doc.parse_status = "processing"
        await db.commit()

        try:
            # 1. Extract text based on file type
            if file_path:
                ext = Path(file_path).suffix.lower()
                if ext == ".pdf":
                    text = extract_text_from_pdf(file_path)
                elif ext in (".xlsx", ".xls"):
                    text = extract_text_from_excel(file_path)
                else:
                    raise ValueError(f"Unsupported file type: {ext}")
            else:
                form_data = doc.parsed_data.get("form_data", {})
                text = json.dumps(form_data) if form_data else ""

            # 2. LLM analysis
            from app.parsers.pitch_deck import MODEL_PATH
            from llama_cpp import Llama
            llm = Llama(model_path=str(MODEL_PATH), n_ctx=2048, n_threads=4, verbose=False)
            prompt = build_analysis_prompt(text, doc.document_type)
            response = llm(prompt, max_tokens=1024, temperature=0.0, stop=["```"])
            raw_output = response["choices"][0]["text"].strip()

            # Store raw output for debugging
            doc.parsed_data = {**(doc.parsed_data or {}), "raw_llm_output": raw_output}

            # 3. Parse based on document type
            if doc.document_type in ("pitch_deck", "funding_application"):
                # JSON path (unchanged)
                if raw_output.startswith("```"):
                    json_str = raw_output.split("```")[1]
                    if json_str.startswith("json"):
                        json_str = json_str[4:]
                else:
                    json_str = raw_output
                try:
                    analysis = json.loads(json_str)
                except json.JSONDecodeError:
                    from json_repair import repair_json
                    repaired = repair_json(json_str)
                    analysis = json.loads(repaired)

                metrics = []
                if analysis.get("projected_revenue_usd"):
                    metrics.append({
                        "metric_name": "projected_revenue",
                        "amount": analysis["projected_revenue_usd"],
                        "currency": "USD",
                        "is_projection": True
                    })
            else:
                # Financial document: line-by-line parsing
                analysis = None
                metrics = parse_financial_lines(raw_output)

            # 4. Store extracted data
            doc.parsed_data = {**(doc.parsed_data or {}), "llm_analysis": analysis}
            doc.confidence_scores = {"overall": 0.85}
            doc.parse_status = "completed"

            # 5. Create financial records
            for metric in metrics:
                db.add(FinancialRecord(
                    company_id=company_id,
                    document_id=doc_id,
                    metric_name=metric["metric_name"],
                    amount=float(metric["amount"]),
                    currency=metric.get("currency", "USD"),
                    usd_equivalent=float(metric["amount"]) if metric.get("currency", "USD") == "USD" else None,
                    period_end=metric.get("period_end"),
                    is_estimate=metric.get("is_projection", False)
                ))

            # 6. If pitch deck / funding application, run full valuation engines
            if doc.document_type in ("pitch_deck", "funding_application") and analysis:
                sector = analysis.get("sector", "Other")
                stage = analysis.get("stage", "Seed")

                baseline_q = await db.execute(
                    select(SectorBaseline.baseline_usd)
                    .where(SectorBaseline.sector == sector, SectorBaseline.stage == stage)
                )
                baseline_val = baseline_q.scalar() or 500_000
                if isinstance(baseline_val, Decimal):
                    baseline_val = float(baseline_val)

                scorecard_ratings = {k: v["rating"] for k, v in analysis["scorecard_ratings"].items()}
                scorecard_result = scorecard_valuation(scorecard_ratings, baseline_valuation=baseline_val)
                db.add(Valuation(
                    company_id=company_id,
                    method="scorecard",
                    calculated_value_usd=scorecard_result["pre_money_usd"],
                    pre_money_usd=scorecard_result["pre_money_usd"],
                    assumptions={
                        "ratings": scorecard_ratings,
                        "weights": scorecard_result["weights"],
                        "baseline": baseline_val,
                        "source": "llm"
                    },
                    output_details=scorecard_result,
                    performed_by="llm"
                ))

                if analysis.get("projected_revenue_usd"):
                    exit_mult_q = await db.execute(
                        select(ExitMultiple.multiple)
                        .where(ExitMultiple.sector == sector, ExitMultiple.stage == stage)
                    )
                    exit_mult = exit_mult_q.scalar() or 3.0
                    if isinstance(exit_mult, Decimal):
                        exit_mult = float(exit_mult)

                    investment = 0
                    if analysis.get("deal_terms") and analysis["deal_terms"].get("funding_requirement_usd"):
                        investment = float(analysis["deal_terms"]["funding_requirement_usd"])

                    vc_result = vc_method(
                        projected_revenue=float(analysis["projected_revenue_usd"]),
                        exit_multiple=exit_mult,
                        target_return=10,
                        investment_amount=investment
                    )
                    db.add(Valuation(
                        company_id=company_id,
                        method="venture_capital",
                        calculated_value_usd=vc_result["pre_money_usd"],
                        pre_money_usd=vc_result["pre_money_usd"],
                        post_money_usd=vc_result["post_money_usd"],
                        assumptions=vc_result["assumptions"],
                        output_details=vc_result,
                        performed_by="llm"
                    ))

                risk_scores = {k: v["score"] for k, v in analysis["risk_scores"].items()}
                risk_result = calculate_risk_score(risk_scores)
                db.add(RiskAssessment(
                    company_id=company_id,
                    overall_score=risk_result["overall_score"],
                    category_scores=risk_result,
                    assessment_notes="Auto‑generated from LLM analysis"
                ))

                rule_q = await db.execute(
                    select(RiskAdjustmentRule.category, RiskAdjustmentRule.score, RiskAdjustmentRule.adjustment_pct)
                )
                rules = rule_q.all()
                adjustment_rules: Dict[str, Dict[int, int]] = {}
                for cat, sc, pct in rules:
                    if cat not in adjustment_rules:
                        adjustment_rules[cat] = {}
                    adjustment_rules[cat][sc] = pct

                risk_adj_result = risk_adjusted_valuation(
                    baseline_val, risk_scores, risk_result["weights"],
                    adjustment_rules=adjustment_rules if adjustment_rules else None
                )
                db.add(Valuation(
                    company_id=company_id,
                    method="risk_adjusted",
                    calculated_value_usd=risk_adj_result["adjusted_valuation_usd"],
                    pre_money_usd=risk_adj_result["adjusted_valuation_usd"],
                    assumptions={
                        "baseline": baseline_val,
                        "risk_scores": risk_scores,
                        "weights": risk_result["weights"],
                        "adjustment_rules_used": adjustment_rules
                    },
                    output_details=risk_adj_result,
                    performed_by="llm"
                ))

                if "valuation_range_llm" in analysis:
                    doc.parsed_data["valuation_range_llm"] = analysis["valuation_range_llm"]

            await db.commit()
        except Exception as e:
            doc.parse_status = "failed"
            doc.parsed_data = {"error": str(e), "raw_llm_output": raw_output}
            await db.commit()


async def start_consumer():
    consumer = AIOKafkaConsumer(
        TOPIC,
        bootstrap_servers=REDPANDA_BROKER,
        group_id="pfurex-parser",
        auto_offset_reset="earliest",
        enable_auto_commit=True,
        max_poll_interval_ms=600000,
        session_timeout_ms=60000,
        heartbeat_interval_ms=20000,
    )
    await consumer.start()
    try:
        async for msg in consumer:
            await process_document(msg)
    finally:
        await consumer.stop()
